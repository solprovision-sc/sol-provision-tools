#!/usr/bin/env python3
"""
Generate the starmap ore-concentration heat-map dataset from the org mining
spreadsheet.

Reads tools/solprovision-mining-tool.xlsx (the LEDGER sheet, which is the source
the HEAT_MAP pivot is built on) and writes a static JS module the starmap loads:

    app/static/starmap/data/ore_heatmap.js

This mirrors the HEAT_MAP pivot: rows = Location, columns = Found Ore, value =
% of that location's finds that were each ore (COUNTA of Found Ore as % of row).

Each Location ("SYSTEM - PARENT - POI") is resolved here to a render *anchor*
(a body name, a Lagrange code, an asteroid belt, or "none" when it can't be
placed) so the runtime only has to look the anchor up against the live scene —
no fragile string parsing in the browser. This is a TRIAL data drop; the eventual
production path swaps the spreadsheet read for a live DB query but keeps the same
output shape.

Run:  python tools/gen_ore_heatmap.py
"""
from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "tools" / "solprovision-mining-tool.xlsx"
OUT = ROOT / "app" / "static" / "starmap" / "data" / "ore_heatmap.js"

# LEDGER column indices (1-based): C = Location, G = Found Ore.
COL_LOCATION = 3
COL_FOUND_ORE = 7


def norm_ws(s: str) -> str:
    """Collapse runs of whitespace so 'GLACIEAN BELT  - Belt ' compares cleanly."""
    return " ".join(str(s).split())


# Location ("SYSTEM - PARENT - POI", whitespace-normalised) -> render anchor.
#   body     : matched by name against world.bodyIndex (planets/moons + belts)
#   lagrange : matched by "ARC-L1" style code against the live Lagrange POIs
#   belt     : an asteroid belt body (disk dropped at the belt's ring position)
#   none     : can't be geolocated from current data -> kept in data, not drawn
#
# A few intentional approximations (no distinct body exists for them yet):
#   'Yela Belt'      -> Yela moon          (Crusader's belt hugs Yela)
#   'GLACIEAN BELT'  -> Glaciem Ring belt  (spelling variant in the ledger)
#   Pyro 'Select Site' rows -> their parent planet
ANCHORS: dict[str, dict] = {
    # ── STANTON ──
    "STANTON - AARON HALO - Belt":    {"kind": "belt", "name": "Aaron Halo"},
    "STANTON - ARCCORP - Arc-L1":     {"kind": "lagrange", "code": "ARC-L1"},
    "STANTON - ARCCORP - Arc-L5":     {"kind": "lagrange", "code": "ARC-L5"},
    "STANTON - ARCCORP - Wala":       {"kind": "body", "name": "Wala"},
    "STANTON - CRUSADER - CRU-L1":    {"kind": "lagrange", "code": "CRU-L1"},
    "STANTON - CRUSADER - CRU-L4":    {"kind": "lagrange", "code": "CRU-L4"},
    "STANTON - CRUSADER - Cellin":    {"kind": "body", "name": "Cellin"},
    "STANTON - CRUSADER - Daymar":    {"kind": "body", "name": "Daymar"},
    "STANTON - CRUSADER - Yela":      {"kind": "body", "name": "Yela"},
    "STANTON - CRUSADER - Yela Belt": {"kind": "body", "name": "Yela"},
    "STANTON - HURSTON - Aberdeen":   {"kind": "body", "name": "Aberdeen"},
    "STANTON - HURSTON - HUR-L3":     {"kind": "lagrange", "code": "HUR-L3"},
    "STANTON - HURSTON - HUR-L4":     {"kind": "lagrange", "code": "HUR-L4"},
    "STANTON - MICROTECH - Euterpe":  {"kind": "body", "name": "Euterpe"},
    "STANTON - MICROTECH - MIC-L1":   {"kind": "lagrange", "code": "MIC-L1"},
    "STANTON - MICROTECH - Mic":      {"kind": "body", "name": "microTech"},
    # ── NYX ──
    "NYX - GLACIEAN BELT - Belt":     {"kind": "belt", "name": "Glaciem Ring"},
    "NYX - KEEGER BELT - Belt":       {"kind": "belt", "name": "Keeger Belt"},
    # ── PYRO (sparse; mining-base codes can't be placed yet) ──
    "PYRO - MINING - RAB-KNAP":         {"kind": "none"},
    "PYRO - MINING BASE - RMB-NIGH":    {"kind": "none"},
    "PYRO - MINING BASE - Select Site": {"kind": "none"},
    "PYRO - PYRO 2 - Select Site":      {"kind": "body", "name": "Monox"},
    "PYRO - PYRO 4 - PY4":              {"kind": "body", "name": "Pyro IV"},
    "PYRO - PYRO 4 - Select Site":      {"kind": "body", "name": "Pyro IV"},
}

SYSTEM_OF = {"STANTON": "stanton", "PYRO": "pyro", "NYX": "nyx"}


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["LEDGER"]

    counts: dict[str, Counter] = defaultdict(Counter)
    for r in range(2, ws.max_row + 1):
        loc = ws.cell(r, COL_LOCATION).value
        ore = ws.cell(r, COL_FOUND_ORE).value
        if not loc or not ore:
            continue
        counts[norm_ws(loc)][str(ore).strip()] += 1

    all_ores = sorted({o for c in counts.values() for o in c})
    locations = []
    max_pct = 0.0
    unmapped = []

    for loc in sorted(counts):
        anchor = ANCHORS.get(loc)
        if anchor is None:
            unmapped.append(loc)
            anchor = {"kind": "none"}
        system = SYSTEM_OF.get(loc.split(" - ", 1)[0], None)
        total = sum(counts[loc].values())
        ores = {ore: round(100 * n / total, 1) for ore, n in counts[loc].items()}
        max_pct = max(max_pct, *ores.values())
        # Strip the leading "SYSTEM - " for a tidier on-map label.
        label = re.sub(r"^[A-Z]+ - ", "", loc)
        locations.append({
            "system": system,
            "label": label,
            "anchor": anchor,
            "samples": total,
            "ores": ores,
        })

    payload = {
        "meta": {
            "ores": all_ores,
            "maxPct": round(max_pct, 1),
            "source": "solprovision-mining-tool.xlsx :: LEDGER (HEAT_MAP pivot)",
        },
        "locations": locations,
    }

    banner = (
        "// ═══════════════════════════════════════════════════════════════════\n"
        "//  ORE CONCENTRATION HEAT-MAP DATA  (auto-generated — do not edit)\n"
        "//  Source: tools/solprovision-mining-tool.xlsx (LEDGER / HEAT_MAP pivot)\n"
        "//  Regenerate: python tools/gen_ore_heatmap.py\n"
        "//\n"
        "//  Per location: % of that location's recorded finds that were each ore\n"
        "//  (matches the workbook's COUNTA-of-Found-Ore as %-of-row). `anchor`\n"
        "//  tells the renderer where to drop the heat disk in the live scene.\n"
        "// ═══════════════════════════════════════════════════════════════════\n\n"
    )
    OUT.write_text(
        banner + "export const ORE_HEATMAP = " + json.dumps(payload, indent=2) + ";\n",
        encoding="utf-8",
    )

    print(f"Wrote {OUT.relative_to(ROOT)}")
    print(f"  {len(locations)} locations, {len(all_ores)} ores, max {max_pct}%")
    if unmapped:
        print("  NOTE: no anchor defined (rendered as 'none'):")
        for u in unmapped:
            print("   -", u)


if __name__ == "__main__":
    main()
